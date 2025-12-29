"""Excel workbook generation for grading matrices."""

from typing import Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, AreaChart, Reference


def col_letter(col_num: int) -> str:
    """Convert 1-based column number to Excel column letter."""
    return get_column_letter(col_num)


def build_qualitative_formula(cell: str, grades: list[dict], scale: dict) -> str:
    """Build a nested IF formula for qualitative grade mapping based on ranges."""
    if not grades:
        return ""
    
    # Sort grades by min score descending (highest first)
    sorted_grades = sorted(grades, key=lambda x: x["min"], reverse=True)
    
    # Build nested IFs: =IF(cell="","",IF(cell>=90,"Excellent",IF(cell>=80,"Good",...)))
    formula = f'=IF({cell}="","",'
    
    for i, grade in enumerate(sorted_grades):
        label = grade["label"]
        min_score = grade["min"]
        
        if i < len(sorted_grades) - 1:
            formula += f'IF({cell}>={min_score},"{label}",'
        else:
            # Last one - just return the label
            formula += f'"{label}"'
    
    # Close all the IFs
    formula += ")" * len(sorted_grades)
    
    return formula


def get_final_grade_column(config: dict) -> int:
    """Calculate which column number contains the Final Grade."""
    set_a_count = len(config["set_a"]["projects"])
    set_b_count = len(config["set_b"]["projects"])
    
    col_final = 1 + set_a_count + 1 + set_b_count + 1 + 1
    return col_final


def create_grade_sheet(
    ws,
    students: list[str],
    config: dict[str, Any],
    grades_data: pd.DataFrame | None = None
):
    """
    Create a grade sheet with headers, student rows, and formulas.
    
    Args:
        ws: Worksheet to populate
        students: List of student names
        config: Configuration dictionary
        grades_data: Optional DataFrame with grades (columns: Student Name, A_Project1, B_Project1, etc.)
    """
    set_a = config["set_a"]
    set_b = config["set_b"]
    scale = config.get("scale", {"min": 0, "max": 100, "decimal_places": 1})
    qualitative_grades = config.get("qualitative_grades", [])
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_fill_a = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_fill_b = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    avg_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    final_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    qual_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Build column structure
    col = 1
    col_student = col
    col += 1
    
    col_set_a_start = col
    col_set_a_end = col + len(set_a["projects"]) - 1
    col = col_set_a_end + 1
    
    col_avg_a = col
    col += 1
    
    col_set_b_start = col
    col_set_b_end = col + len(set_b["projects"]) - 1
    col = col_set_b_end + 1
    
    col_avg_b = col
    col += 1
    
    col_final = col
    col += 1
    
    col_qual = col
    
    # --- Row 1: Main headers ---
    row = 1
    
    ws.cell(row=row, column=col_student, value="Student Name")
    ws.cell(row=row, column=col_student).font = header_font
    ws.cell(row=row, column=col_student).fill = header_fill
    ws.cell(row=row, column=col_student).alignment = center_align
    ws.cell(row=row, column=col_student).border = thin_border
    
    # Set A header (merged)
    weight_a = int(set_a["weight"] * 100)
    ws.cell(row=row, column=col_set_a_start, value=f"{set_a['name']} ({weight_a}%)")
    for c in range(col_set_a_start, col_avg_a):
        ws.cell(row=row, column=c).font = header_font
        ws.cell(row=row, column=c).fill = header_fill_a
        ws.cell(row=row, column=c).alignment = center_align
        ws.cell(row=row, column=c).border = thin_border
    if len(set_a["projects"]) > 1:
        ws.merge_cells(start_row=row, start_column=col_set_a_start, end_row=row, end_column=col_set_a_end)
    
    ws.cell(row=row, column=col_avg_a, value="Avg A")
    ws.cell(row=row, column=col_avg_a).font = header_font
    ws.cell(row=row, column=col_avg_a).fill = avg_fill
    ws.cell(row=row, column=col_avg_a).font = Font(bold=True)
    ws.cell(row=row, column=col_avg_a).alignment = center_align
    ws.cell(row=row, column=col_avg_a).border = thin_border
    
    # Set B header (merged)
    weight_b = int(set_b["weight"] * 100)
    ws.cell(row=row, column=col_set_b_start, value=f"{set_b['name']} ({weight_b}%)")
    for c in range(col_set_b_start, col_avg_b):
        ws.cell(row=row, column=c).font = header_font
        ws.cell(row=row, column=c).fill = header_fill_b
        ws.cell(row=row, column=c).alignment = center_align
        ws.cell(row=row, column=c).border = thin_border
    if len(set_b["projects"]) > 1:
        ws.merge_cells(start_row=row, start_column=col_set_b_start, end_row=row, end_column=col_set_b_end)
    
    ws.cell(row=row, column=col_avg_b, value="Avg B")
    ws.cell(row=row, column=col_avg_b).font = header_font
    ws.cell(row=row, column=col_avg_b).fill = avg_fill
    ws.cell(row=row, column=col_avg_b).font = Font(bold=True)
    ws.cell(row=row, column=col_avg_b).alignment = center_align
    ws.cell(row=row, column=col_avg_b).border = thin_border
    
    ws.cell(row=row, column=col_final, value="Final Grade")
    ws.cell(row=row, column=col_final).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=col_final).fill = final_fill
    ws.cell(row=row, column=col_final).font = Font(bold=True)
    ws.cell(row=row, column=col_final).alignment = center_align
    ws.cell(row=row, column=col_final).border = thin_border
    
    ws.cell(row=row, column=col_qual, value="Qualitative")
    ws.cell(row=row, column=col_qual).font = Font(bold=True)
    ws.cell(row=row, column=col_qual).fill = qual_fill
    ws.cell(row=row, column=col_qual).alignment = center_align
    ws.cell(row=row, column=col_qual).border = thin_border
    
    # --- Row 2: Project names ---
    row = 2
    
    ws.cell(row=row, column=col_student, value="")
    ws.cell(row=row, column=col_student).border = thin_border
    
    for i, proj in enumerate(set_a["projects"]):
        c = col_set_a_start + i
        ws.cell(row=row, column=c, value=proj)
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).alignment = center_align
        ws.cell(row=row, column=c).border = thin_border
    
    ws.cell(row=row, column=col_avg_a, value="")
    ws.cell(row=row, column=col_avg_a).border = thin_border
    
    for i, proj in enumerate(set_b["projects"]):
        c = col_set_b_start + i
        ws.cell(row=row, column=c, value=proj)
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).alignment = center_align
        ws.cell(row=row, column=c).border = thin_border
    
    ws.cell(row=row, column=col_avg_b, value="")
    ws.cell(row=row, column=col_avg_b).border = thin_border
    ws.cell(row=row, column=col_final, value="")
    ws.cell(row=row, column=col_final).border = thin_border
    ws.cell(row=row, column=col_qual, value="")
    ws.cell(row=row, column=col_qual).border = thin_border
    
    # --- Student rows with formulas ---
    for student_idx, student_name in enumerate(students):
        row = 3 + student_idx
        
        # Student name
        ws.cell(row=row, column=col_student, value=student_name)
        ws.cell(row=row, column=col_student).border = thin_border
        
        # Set A project cells
        for i, proj in enumerate(set_a["projects"]):
            c = col_set_a_start + i
            col_name = f"A_{proj}"
            
            # Get grade value if available
            grade_value = ""
            if grades_data is not None and col_name in grades_data.columns:
                student_row = grades_data[grades_data["Student Name"] == student_name]
                if not student_row.empty:
                    val = student_row[col_name].iloc[0]
                    if pd.notna(val) and val != "":
                        grade_value = val
            
            ws.cell(row=row, column=c, value=grade_value)
            ws.cell(row=row, column=c).alignment = center_align
            ws.cell(row=row, column=c).border = thin_border
        
        # Avg A formula
        range_a_start = col_letter(col_set_a_start)
        range_a_end = col_letter(col_set_a_end)
        avg_a_formula = f"=IF(COUNT({range_a_start}{row}:{range_a_end}{row})>0,AVERAGE({range_a_start}{row}:{range_a_end}{row}),\"\")"
        ws.cell(row=row, column=col_avg_a, value=avg_a_formula)
        ws.cell(row=row, column=col_avg_a).alignment = center_align
        ws.cell(row=row, column=col_avg_a).fill = avg_fill
        ws.cell(row=row, column=col_avg_a).border = thin_border
        
        # Set B project cells
        for i, proj in enumerate(set_b["projects"]):
            c = col_set_b_start + i
            col_name = f"B_{proj}"
            
            # Get grade value if available
            grade_value = ""
            if grades_data is not None and col_name in grades_data.columns:
                student_row = grades_data[grades_data["Student Name"] == student_name]
                if not student_row.empty:
                    val = student_row[col_name].iloc[0]
                    if pd.notna(val) and val != "":
                        grade_value = val
            
            ws.cell(row=row, column=c, value=grade_value)
            ws.cell(row=row, column=c).alignment = center_align
            ws.cell(row=row, column=c).border = thin_border
        
        # Avg B formula
        range_b_start = col_letter(col_set_b_start)
        range_b_end = col_letter(col_set_b_end)
        avg_b_formula = f"=IF(COUNT({range_b_start}{row}:{range_b_end}{row})>0,AVERAGE({range_b_start}{row}:{range_b_end}{row}),\"\")"
        ws.cell(row=row, column=col_avg_b, value=avg_b_formula)
        ws.cell(row=row, column=col_avg_b).alignment = center_align
        ws.cell(row=row, column=col_avg_b).fill = avg_fill
        ws.cell(row=row, column=col_avg_b).border = thin_border
        
        # Final grade formula
        avg_a_cell = f"{col_letter(col_avg_a)}{row}"
        avg_b_cell = f"{col_letter(col_avg_b)}{row}"
        weight_a = set_a["weight"]
        weight_b = set_b["weight"]
        final_formula = f'=IF(AND(ISNUMBER({avg_a_cell}),ISNUMBER({avg_b_cell})),{avg_a_cell}*{weight_a}+{avg_b_cell}*{weight_b},IF(ISNUMBER({avg_a_cell}),{avg_a_cell}*{weight_a},IF(ISNUMBER({avg_b_cell}),{avg_b_cell}*{weight_b},"")))'
        ws.cell(row=row, column=col_final, value=final_formula)
        ws.cell(row=row, column=col_final).alignment = center_align
        ws.cell(row=row, column=col_final).fill = final_fill
        ws.cell(row=row, column=col_final).border = thin_border
        
        # Qualitative grade formula
        final_cell = f"{col_letter(col_final)}{row}"
        qual_formula = build_qualitative_formula(final_cell, qualitative_grades, scale)
        ws.cell(row=row, column=col_qual, value=qual_formula)
        ws.cell(row=row, column=col_qual).alignment = center_align
        ws.cell(row=row, column=col_qual).fill = qual_fill
        ws.cell(row=row, column=col_qual).border = thin_border
    
    # Adjust column widths
    ws.column_dimensions[col_letter(col_student)].width = 25
    for c in range(col_set_a_start, col_set_a_end + 1):
        ws.column_dimensions[col_letter(c)].width = 12
    ws.column_dimensions[col_letter(col_avg_a)].width = 10
    for c in range(col_set_b_start, col_set_b_end + 1):
        ws.column_dimensions[col_letter(c)].width = 12
    ws.column_dimensions[col_letter(col_avg_b)].width = 10
    ws.column_dimensions[col_letter(col_final)].width = 12
    ws.column_dimensions[col_letter(col_qual)].width = 20


def create_total_sheet(
    ws,
    students: list[str],
    config: dict[str, Any]
):
    """Create a Total sheet that summarizes all trimesters with year average."""
    trimesters = config["trimesters"]
    qualitative_grades = config.get("qualitative_grades", [])
    scale = config.get("scale", {"min": 0, "max": 100, "decimal_places": 1})
    
    # Get the column positions for Final Grade and Qualitative in trimester sheets
    col_final_in_trimester = get_final_grade_column(config)
    col_qual_in_trimester = col_final_in_trimester + 1
    final_col_letter_str = col_letter(col_final_in_trimester)
    qual_col_letter_str = col_letter(col_qual_in_trimester)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    trimester_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    qual_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    year_avg_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    year_qual_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Column structure
    col = 1
    col_student = col
    col += 1
    
    trimester_cols = []
    for _ in trimesters:
        trimester_cols.append({"grade": col, "qual": col + 1})
        col += 2
    
    col_year_avg = col
    col += 1
    col_year_qual = col
    
    # --- Row 1: Headers ---
    row = 1
    
    ws.cell(row=row, column=col_student, value="Student Name")
    ws.cell(row=row, column=col_student).font = header_font
    ws.cell(row=row, column=col_student).fill = header_fill
    ws.cell(row=row, column=col_student).alignment = center_align
    ws.cell(row=row, column=col_student).border = thin_border
    
    for i, trimester in enumerate(trimesters):
        ws.cell(row=row, column=trimester_cols[i]["grade"], value=f"{trimester} Grade")
        ws.cell(row=row, column=trimester_cols[i]["grade"]).font = Font(bold=True)
        ws.cell(row=row, column=trimester_cols[i]["grade"]).fill = trimester_fill
        ws.cell(row=row, column=trimester_cols[i]["grade"]).alignment = center_align
        ws.cell(row=row, column=trimester_cols[i]["grade"]).border = thin_border
        
        ws.cell(row=row, column=trimester_cols[i]["qual"], value=f"{trimester} Qual")
        ws.cell(row=row, column=trimester_cols[i]["qual"]).font = Font(bold=True)
        ws.cell(row=row, column=trimester_cols[i]["qual"]).fill = qual_fill
        ws.cell(row=row, column=trimester_cols[i]["qual"]).alignment = center_align
        ws.cell(row=row, column=trimester_cols[i]["qual"]).border = thin_border
    
    ws.cell(row=row, column=col_year_avg, value="Year Average")
    ws.cell(row=row, column=col_year_avg).font = Font(bold=True)
    ws.cell(row=row, column=col_year_avg).fill = year_avg_fill
    ws.cell(row=row, column=col_year_avg).alignment = center_align
    ws.cell(row=row, column=col_year_avg).border = thin_border
    
    ws.cell(row=row, column=col_year_qual, value="Year Qualitative")
    ws.cell(row=row, column=col_year_qual).font = Font(bold=True)
    ws.cell(row=row, column=col_year_qual).fill = year_qual_fill
    ws.cell(row=row, column=col_year_qual).alignment = center_align
    ws.cell(row=row, column=col_year_qual).border = thin_border
    
    # --- Student rows ---
    for student_idx, student_name in enumerate(students):
        row = 2 + student_idx
        data_row_in_trimester = 3 + student_idx
        
        ws.cell(row=row, column=col_student, value=student_name)
        ws.cell(row=row, column=col_student).border = thin_border
        
        grade_cells = []
        for i, trimester in enumerate(trimesters):
            grade_ref = f"='{trimester}'!{final_col_letter_str}{data_row_in_trimester}"
            ws.cell(row=row, column=trimester_cols[i]["grade"], value=grade_ref)
            ws.cell(row=row, column=trimester_cols[i]["grade"]).alignment = center_align
            ws.cell(row=row, column=trimester_cols[i]["grade"]).fill = trimester_fill
            ws.cell(row=row, column=trimester_cols[i]["grade"]).border = thin_border
            
            grade_cells.append(f"{col_letter(trimester_cols[i]['grade'])}{row}")
            
            qual_ref = f"='{trimester}'!{qual_col_letter_str}{data_row_in_trimester}"
            ws.cell(row=row, column=trimester_cols[i]["qual"], value=qual_ref)
            ws.cell(row=row, column=trimester_cols[i]["qual"]).alignment = center_align
            ws.cell(row=row, column=trimester_cols[i]["qual"]).fill = qual_fill
            ws.cell(row=row, column=trimester_cols[i]["qual"]).border = thin_border
        
        grade_cells_str = ",".join(grade_cells)
        year_avg_formula = f"=IF(COUNT({grade_cells_str})>0,AVERAGE({grade_cells_str}),\"\")"
        ws.cell(row=row, column=col_year_avg, value=year_avg_formula)
        ws.cell(row=row, column=col_year_avg).alignment = center_align
        ws.cell(row=row, column=col_year_avg).fill = year_avg_fill
        ws.cell(row=row, column=col_year_avg).border = thin_border
        
        year_avg_cell = f"{col_letter(col_year_avg)}{row}"
        year_qual_formula = build_qualitative_formula(year_avg_cell, qualitative_grades, scale)
        ws.cell(row=row, column=col_year_qual, value=year_qual_formula)
        ws.cell(row=row, column=col_year_qual).alignment = center_align
        ws.cell(row=row, column=col_year_qual).fill = year_qual_fill
        ws.cell(row=row, column=col_year_qual).border = thin_border
    
    # Adjust column widths
    ws.column_dimensions[col_letter(col_student)].width = 25
    for i in range(len(trimesters)):
        ws.column_dimensions[col_letter(trimester_cols[i]["grade"])].width = 14
        ws.column_dimensions[col_letter(trimester_cols[i]["qual"])].width = 18
    ws.column_dimensions[col_letter(col_year_avg)].width = 14
    ws.column_dimensions[col_letter(col_year_qual)].width = 18


def create_chart_sheet(
    ws,
    chart_data: pd.DataFrame,
    chart_config: dict[str, Any]
):
    """
    Create a sheet with chart data and embedded chart.
    
    Args:
        ws: Worksheet to populate
        chart_data: DataFrame with Student column and data columns
        chart_config: Dict with 'columns' and 'chart_type' keys
    """
    if chart_data is None or chart_data.empty:
        ws["A1"] = "No chart data available"
        return
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Write headers
    for col_idx, col_name in enumerate(chart_data.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Write data
    for row_idx, row in enumerate(chart_data.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border
    
    # Adjust column widths
    for col_idx, col_name in enumerate(chart_data.columns, 1):
        ws.column_dimensions[col_letter(col_idx)].width = max(15, len(str(col_name)) + 2)
    
    # Create chart
    chart_type = chart_config.get("chart_type", "bar_chart")
    
    if chart_type == "line_chart":
        chart = LineChart()
    elif chart_type == "area_chart":
        chart = AreaChart()
    else:
        chart = BarChart()
    
    chart.title = "Grade Comparison"
    chart.style = 10
    chart.x_axis.title = "Students"
    chart.y_axis.title = "Grade"
    
    # Data reference (skip first column which is Student names)
    num_rows = len(chart_data) + 1  # +1 for header
    num_cols = len(chart_data.columns)
    
    if num_cols > 1:
        # Data series
        data = Reference(ws, min_col=2, min_row=1, max_col=num_cols, max_row=num_rows)
        # Categories (student names)
        cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        
        # Chart size
        chart.width = 20
        chart.height = 12
        
        # Place chart below the data
        chart_position = f"A{num_rows + 3}"
        ws.add_chart(chart, chart_position)


def generate_workbook(
    config: dict[str, Any],
    students: list[str],
    grades_by_trimester: dict[str, pd.DataFrame] | None = None,
    chart_data: pd.DataFrame | None = None,
    chart_config: dict[str, Any] | None = None
) -> Workbook:
    """
    Generate the Excel workbook with all trimester sheets.
    
    Args:
        config: Configuration dictionary
        students: List of student names
        grades_by_trimester: Optional dict mapping trimester name to DataFrame with grades
        chart_data: Optional DataFrame with chart data
        chart_config: Optional dict with chart configuration
        
    Returns:
        openpyxl Workbook object
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create a sheet for each trimester
    for trimester in config["trimesters"]:
        ws = wb.create_sheet(title=trimester)
        grades_data = None
        if grades_by_trimester and trimester in grades_by_trimester:
            grades_data = grades_by_trimester[trimester]
        create_grade_sheet(ws, students, config, grades_data)
    
    # Create the Total sheet
    ws_total = wb.create_sheet(title="Total")
    create_total_sheet(ws_total, students, config)
    
    # Create chart sheet if chart data is provided
    if chart_data is not None and not chart_data.empty and chart_config:
        ws_chart = wb.create_sheet(title="Charts")
        create_chart_sheet(ws_chart, chart_data, chart_config)
    
    # Create instructions sheet
    ws_info = wb.create_sheet(title="Instructions", index=0)
    ws_info["A1"] = "Student Grading Matrix"
    ws_info["A1"].font = Font(bold=True, size=16)
    ws_info["A3"] = "How to use:"
    ws_info["A3"].font = Font(bold=True)
    ws_info["A4"] = f"1. Go to each Trimester tab and enter grades ({config['scale']['min']}-{config['scale']['max']})"
    ws_info["A5"] = "2. Averages, Final Grade, and Qualitative calculate automatically"
    ws_info["A6"] = "3. The Total tab shows all trimester summaries and year average"
    ws_info["A8"] = "Configuration:"
    ws_info["A8"].font = Font(bold=True)
    
    row = 9
    ws_info[f"A{row}"] = f"Grade Scale: {config['scale']['min']} - {config['scale']['max']}"
    row += 1
    ws_info[f"A{row}"] = f"Set A Weight: {int(config['set_a']['weight'] * 100)}%"
    row += 1
    ws_info[f"A{row}"] = f"Set A Projects: {', '.join(config['set_a']['projects'])}"
    row += 1
    ws_info[f"A{row}"] = f"Set B Weight: {int(config['set_b']['weight'] * 100)}%"
    row += 1
    ws_info[f"A{row}"] = f"Set B Projects: {', '.join(config['set_b']['projects'])}"
    row += 2
    
    qualitative_grades = config.get("qualitative_grades", [])
    if qualitative_grades:
        ws_info[f"A{row}"] = "Qualitative Grade Ranges:"
        ws_info[f"A{row}"].font = Font(bold=True)
        row += 1
        for grade in qualitative_grades:
            ws_info[f"A{row}"] = f"  {grade['label']}: {grade['min']} - {grade['max']}"
            row += 1
    
    ws_info.column_dimensions["A"].width = 50
    
    return wb

