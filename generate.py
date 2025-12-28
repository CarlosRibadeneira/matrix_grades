#!/usr/bin/env python3
"""
Student Grading Matrix Generator

Reads configuration from config.json and student names from students.txt,
then generates an Excel file with grade sheets for each trimester.

Usage:
    1. Edit config.json to set up projects, weights, and qualitative grades
    2. Paste student names into students.txt (one per line)
    3. Run: python generate.py
    4. Open the generated Excel file and enter grades
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def load_config(config_path: str = "config.json") -> dict:
    """Load configuration from JSON file."""
    with open(config_path, "r") as f:
        return json.load(f)


def load_students(students_path: str = "students.txt") -> list[str]:
    """Load student names from text file, ignoring comments and empty lines."""
    students = []
    with open(students_path, "r") as f:
        for line in f:
            line = line.strip()
            # Skip empty lines and comments
            if line and not line.startswith("#"):
                students.append(line)
    return students


def col_letter(col_num: int) -> str:
    """Convert 1-based column number to Excel column letter."""
    return get_column_letter(col_num)


def build_qualitative_formula(cell: str, grades: list[dict]) -> str:
    """Build a nested IF formula for qualitative grade mapping based on ranges."""
    if not grades:
        return ""
    
    # Sort grades by min score descending (highest first)
    sorted_grades = sorted(grades, key=lambda x: x["min"], reverse=True)
    
    # Build nested IFs: =IF(cell="","",IF(cell>=90,"Excellent",IF(cell>=80,"Good",...)))
    formula = f'=IF({cell}="","",'
    
    for i, grade in enumerate(sorted_grades):
        label = grade["label"]
        min_score = grade["min"]
        
        if i < len(sorted_grades) - 1:
            formula += f'IF({cell}>={min_score},"{label}",'
        else:
            # Last one - just return the label
            formula += f'"{label}"'
    
    # Close all the IFs
    formula += ")" * len(sorted_grades)
    
    return formula


def create_grade_sheet(ws, students: list[str], config: dict):
    """Create a grade sheet with headers, student rows, and formulas."""
    
    set_a = config["set_a"]
    set_b = config["set_b"]
    qualitative_grades = config.get("qualitative_grades", [])
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_fill_a = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_fill_b = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    avg_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    final_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    qual_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Build column structure
    col = 1
    col_student = col
    col += 1
    
    col_set_a_start = col
    col_set_a_end = col + len(set_a["projects"]) - 1
    col = col_set_a_end + 1
    
    col_avg_a = col
    col += 1
    
    col_set_b_start = col
    col_set_b_end = col + len(set_b["projects"]) - 1
    col = col_set_b_end + 1
    
    col_avg_b = col
    col += 1
    
    col_final = col
    col += 1
    
    col_qual = col
    
    # --- Row 1: Main headers ---
    row = 1
    
    ws.cell(row=row, column=col_student, value="Student Name")
    ws.cell(row=row, column=col_student).font = header_font
    ws.cell(row=row, column=col_student).fill = header_fill
    ws.cell(row=row, column=col_student).alignment = center_align
    ws.cell(row=row, column=col_student).border = thin_border
    
    # Set A header (merged)
    weight_a = int(set_a["weight"] * 100)
    ws.cell(row=row, column=col_set_a_start, value=f"{set_a['name']} ({weight_a}%)")
    for c in range(col_set_a_start, col_avg_a):
        ws.cell(row=row, column=c).font = header_font
        ws.cell(row=row, column=c).fill = header_fill_a
        ws.cell(row=row, column=c).alignment = center_align
        ws.cell(row=row, column=c).border = thin_border
    if len(set_a["projects"]) > 1:
        ws.merge_cells(start_row=row, start_column=col_set_a_start, end_row=row, end_column=col_set_a_end)
    
    ws.cell(row=row, column=col_avg_a, value="Avg A")
    ws.cell(row=row, column=col_avg_a).font = header_font
    ws.cell(row=row, column=col_avg_a).fill = avg_fill
    ws.cell(row=row, column=col_avg_a).font = Font(bold=True)
    ws.cell(row=row, column=col_avg_a).alignment = center_align
    ws.cell(row=row, column=col_avg_a).border = thin_border
    
    # Set B header (merged)
    weight_b = int(set_b["weight"] * 100)
    ws.cell(row=row, column=col_set_b_start, value=f"{set_b['name']} ({weight_b}%)")
    for c in range(col_set_b_start, col_avg_b):
        ws.cell(row=row, column=c).font = header_font
        ws.cell(row=row, column=c).fill = header_fill_b
        ws.cell(row=row, column=c).alignment = center_align
        ws.cell(row=row, column=c).border = thin_border
    if len(set_b["projects"]) > 1:
        ws.merge_cells(start_row=row, start_column=col_set_b_start, end_row=row, end_column=col_set_b_end)
    
    ws.cell(row=row, column=col_avg_b, value="Avg B")
    ws.cell(row=row, column=col_avg_b).font = header_font
    ws.cell(row=row, column=col_avg_b).fill = avg_fill
    ws.cell(row=row, column=col_avg_b).font = Font(bold=True)
    ws.cell(row=row, column=col_avg_b).alignment = center_align
    ws.cell(row=row, column=col_avg_b).border = thin_border
    
    ws.cell(row=row, column=col_final, value="Final Grade")
    ws.cell(row=row, column=col_final).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=col_final).fill = final_fill
    ws.cell(row=row, column=col_final).font = Font(bold=True)
    ws.cell(row=row, column=col_final).alignment = center_align
    ws.cell(row=row, column=col_final).border = thin_border
    
    ws.cell(row=row, column=col_qual, value="Qualitative")
    ws.cell(row=row, column=col_qual).font = Font(bold=True)
    ws.cell(row=row, column=col_qual).fill = qual_fill
    ws.cell(row=row, column=col_qual).alignment = center_align
    ws.cell(row=row, column=col_qual).border = thin_border
    
    # --- Row 2: Project names ---
    row = 2
    
    ws.cell(row=row, column=col_student, value="")
    ws.cell(row=row, column=col_student).border = thin_border
    
    for i, proj in enumerate(set_a["projects"]):
        c = col_set_a_start + i
        ws.cell(row=row, column=c, value=proj)
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).alignment = center_align
        ws.cell(row=row, column=c).border = thin_border
    
    ws.cell(row=row, column=col_avg_a, value="")
    ws.cell(row=row, column=col_avg_a).border = thin_border
    
    for i, proj in enumerate(set_b["projects"]):
        c = col_set_b_start + i
        ws.cell(row=row, column=c, value=proj)
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).alignment = center_align
        ws.cell(row=row, column=c).border = thin_border
    
    ws.cell(row=row, column=col_avg_b, value="")
    ws.cell(row=row, column=col_avg_b).border = thin_border
    ws.cell(row=row, column=col_final, value="")
    ws.cell(row=row, column=col_final).border = thin_border
    ws.cell(row=row, column=col_qual, value="")
    ws.cell(row=row, column=col_qual).border = thin_border
    
    # --- Student rows with formulas ---
    for student_idx, student_name in enumerate(students):
        row = 3 + student_idx
        
        # Student name
        ws.cell(row=row, column=col_student, value=student_name)
        ws.cell(row=row, column=col_student).border = thin_border
        
        # Set A project cells (empty for grade entry)
        for i in range(len(set_a["projects"])):
            c = col_set_a_start + i
            ws.cell(row=row, column=c, value="")
            ws.cell(row=row, column=c).alignment = center_align
            ws.cell(row=row, column=c).border = thin_border
        
        # Avg A formula
        range_a_start = col_letter(col_set_a_start)
        range_a_end = col_letter(col_set_a_end)
        avg_a_formula = f"=IF(COUNT({range_a_start}{row}:{range_a_end}{row})>0,AVERAGE({range_a_start}{row}:{range_a_end}{row}),\"\")"
        ws.cell(row=row, column=col_avg_a, value=avg_a_formula)
        ws.cell(row=row, column=col_avg_a).alignment = center_align
        ws.cell(row=row, column=col_avg_a).fill = avg_fill
        ws.cell(row=row, column=col_avg_a).border = thin_border
        
        # Set B project cells (empty for grade entry)
        for i in range(len(set_b["projects"])):
            c = col_set_b_start + i
            ws.cell(row=row, column=c, value="")
            ws.cell(row=row, column=c).alignment = center_align
            ws.cell(row=row, column=c).border = thin_border
        
        # Avg B formula
        range_b_start = col_letter(col_set_b_start)
        range_b_end = col_letter(col_set_b_end)
        avg_b_formula = f"=IF(COUNT({range_b_start}{row}:{range_b_end}{row})>0,AVERAGE({range_b_start}{row}:{range_b_end}{row}),\"\")"
        ws.cell(row=row, column=col_avg_b, value=avg_b_formula)
        ws.cell(row=row, column=col_avg_b).alignment = center_align
        ws.cell(row=row, column=col_avg_b).fill = avg_fill
        ws.cell(row=row, column=col_avg_b).border = thin_border
        
        # Final grade formula
        avg_a_cell = f"{col_letter(col_avg_a)}{row}"
        avg_b_cell = f"{col_letter(col_avg_b)}{row}"
        weight_a = set_a["weight"]
        weight_b = set_b["weight"]
        final_formula = f'=IF(AND(ISNUMBER({avg_a_cell}),ISNUMBER({avg_b_cell})),{avg_a_cell}*{weight_a}+{avg_b_cell}*{weight_b},IF(ISNUMBER({avg_a_cell}),{avg_a_cell}*{weight_a},IF(ISNUMBER({avg_b_cell}),{avg_b_cell}*{weight_b},"")))'
        ws.cell(row=row, column=col_final, value=final_formula)
        ws.cell(row=row, column=col_final).alignment = center_align
        ws.cell(row=row, column=col_final).fill = final_fill
        ws.cell(row=row, column=col_final).border = thin_border
        
        # Qualitative grade formula (auto-calculated from final grade)
        final_cell = f"{col_letter(col_final)}{row}"
        qual_formula = build_qualitative_formula(final_cell, qualitative_grades)
        ws.cell(row=row, column=col_qual, value=qual_formula)
        ws.cell(row=row, column=col_qual).alignment = center_align
        ws.cell(row=row, column=col_qual).fill = qual_fill
        ws.cell(row=row, column=col_qual).border = thin_border
    
    # Adjust column widths
    ws.column_dimensions[col_letter(col_student)].width = 25
    for c in range(col_set_a_start, col_set_a_end + 1):
        ws.column_dimensions[col_letter(c)].width = 12
    ws.column_dimensions[col_letter(col_avg_a)].width = 10
    for c in range(col_set_b_start, col_set_b_end + 1):
        ws.column_dimensions[col_letter(c)].width = 12
    ws.column_dimensions[col_letter(col_avg_b)].width = 10
    ws.column_dimensions[col_letter(col_final)].width = 12
    ws.column_dimensions[col_letter(col_qual)].width = 20


def generate_workbook(config: dict, students: list[str]) -> Workbook:
    """Generate the Excel workbook with all trimester sheets."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create a sheet for each trimester
    for trimester in config["trimesters"]:
        ws = wb.create_sheet(title=trimester)
        create_grade_sheet(ws, students, config)
    
    # Create a legend/instructions sheet
    ws_info = wb.create_sheet(title="Instructions", index=0)
    ws_info["A1"] = "Student Grading Matrix"
    ws_info["A1"].font = Font(bold=True, size=16)
    ws_info["A3"] = "How to use:"
    ws_info["A3"].font = Font(bold=True)
    ws_info["A4"] = "1. Go to each Trimester tab"
    ws_info["A5"] = "2. Enter grades (0-100) in the project columns"
    ws_info["A6"] = "3. Averages, Final Grade, and Qualitative calculate automatically"
    ws_info["A8"] = "Configuration:"
    ws_info["A8"].font = Font(bold=True)
    
    row = 9
    ws_info[f"A{row}"] = f"Set A Weight: {int(config['set_a']['weight'] * 100)}%"
    row += 1
    ws_info[f"A{row}"] = f"Set A Projects: {', '.join(config['set_a']['projects'])}"
    row += 1
    ws_info[f"A{row}"] = f"Set B Weight: {int(config['set_b']['weight'] * 100)}%"
    row += 1
    ws_info[f"A{row}"] = f"Set B Projects: {', '.join(config['set_b']['projects'])}"
    row += 2
    
    qualitative_grades = config.get("qualitative_grades", [])
    if qualitative_grades:
        ws_info[f"A{row}"] = "Qualitative Grade Ranges:"
        ws_info[f"A{row}"].font = Font(bold=True)
        row += 1
        for grade in qualitative_grades:
            ws_info[f"A{row}"] = f"  {grade['label']}: {grade['min']} - {grade['max']}"
            row += 1
    
    ws_info.column_dimensions["A"].width = 50
    
    return wb


def main():
    """Main entry point."""
    print("📊 Student Grading Matrix Generator")
    print("=" * 40)
    
    # Load configuration
    config_path = Path("config.json")
    if not config_path.exists():
        print("❌ Error: config.json not found!")
        print("   Please create a config.json file with your settings.")
        return
    
    config = load_config(config_path)
    print(f"✓ Loaded configuration from {config_path}")
    
    # Load students
    students_path = Path("students.txt")
    if not students_path.exists():
        print("❌ Error: students.txt not found!")
        print("   Please create a students.txt file with student names (one per line).")
        return
    
    students = load_students(students_path)
    if not students:
        print("❌ Error: No students found in students.txt!")
        print("   Add student names (one per line, lines starting with # are ignored).")
        return
    
    print(f"✓ Loaded {len(students)} students from {students_path}")
    
    # Generate workbook
    print("\n📝 Generating Excel file...")
    wb = generate_workbook(config, students)
    
    # Save
    output_file = config.get("output_file", "grades.xlsx")
    wb.save(output_file)
    print(f"✓ Saved to {output_file}")
    
    # Summary
    print("\n" + "=" * 40)
    print("📋 Summary:")
    print(f"   Students: {len(students)}")
    print(f"   Trimesters: {len(config['trimesters'])}")
    print(f"   Set A Projects: {len(config['set_a']['projects'])} ({int(config['set_a']['weight']*100)}% weight)")
    print(f"   Set B Projects: {len(config['set_b']['projects'])} ({int(config['set_b']['weight']*100)}% weight)")
    
    qualitative_grades = config.get("qualitative_grades", [])
    if qualitative_grades:
        print(f"   Qualitative Ranges: {len(qualitative_grades)} levels")
    
    print(f"\n🎉 Open {output_file} and start entering grades!")


if __name__ == "__main__":
    main()
